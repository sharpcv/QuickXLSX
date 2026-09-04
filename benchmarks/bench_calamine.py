#!/usr/bin/env python3
"""Python comparison worker: calamine engine for reads, openpyxl write_only for writes.

python-calamine exposes an openpyxl-compatible read-only API backed by the Rust
calamine engine. Empty cells are returned as '' (never None), so stored-cell
counts exclude both None and the empty string. Writes use openpyxl's write_only
(streaming) mode, which mirrors the C++ streaming writers.
"""
import argparse
import pathlib
import sys

import openpyxl
from python_calamine import CalamineWorkbook


def read(input_path: pathlib.Path) -> tuple[int, int]:
    workbook = CalamineWorkbook.from_path(str(input_path))
    rows = 0
    cells = 0
    for name in workbook.sheet_names:
        sheet = workbook.get_sheet_by_name(name)
        for row in sheet.iter_rows():
            rows += 1
            cells += sum(1 for value in row if value not in (None, ""))
    workbook.close()
    return rows, cells


def filter_write(input_path: pathlib.Path, output_path: pathlib.Path) -> int:
    workbook = CalamineWorkbook.from_path(str(input_path))
    sheet = workbook.get_sheet_by_name(workbook.sheet_names[0])

    output = openpyxl.Workbook(write_only=True)
    output_sheet = output.create_sheet(sheet.name)

    match_col: int | None = None
    matches = 0
    for row in sheet.iter_rows():
        if match_col is None:
            for index, value in enumerate(row):
                if value == "结算客户名称":
                    match_col = index
                    break
            output_sheet.append(list(row))
            continue
        value = row[match_col]
        if value and "西子电商" in str(value):
            output_sheet.append(list(row))
            matches += 1

    output.save(output_path)
    output.close()
    workbook.close()
    return matches


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("mode", choices=("read", "filterwrite"))
    parser.add_argument("input", type=pathlib.Path)
    parser.add_argument("output", nargs="?", type=pathlib.Path)
    args = parser.parse_args()
    if (args.mode == "filterwrite") != (args.output is not None):
        parser.error("filterwrite requires OUTPUT; read does not accept it")

    if args.mode == "read":
        rows, cells = read(args.input)
        print(f"rows={rows} cells={cells}")
        return int(rows == 0 or cells == 0)

    matches = filter_write(args.input, args.output)
    print(f"matches={matches}")
    return int(matches == 0)


if __name__ == "__main__":
    sys.exit(main())
